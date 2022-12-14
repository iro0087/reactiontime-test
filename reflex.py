import time

import matplotlib.pyplot as plt

import sys

import random

import keyboard

from termcolor import colored

from openpyxl import load_workbook, Workbook

doss = load_workbook("reflex.xlsx")

doss1 = doss.active

numd = doss1.cell(row=1, column=2).value + 1

doss1.cell(row=1, column=2).value = numd

namesaved = "refllex" + str(numd) + ".jpg"

t = 1

l = []

rl = []

colorl = ["red", "blue", "yellow", "green"]

while t < 6:

    l.append(t)

    t += 1

t = 0

doss1.cell(row=1, column=1).value = 1

def fun():
    
    t_print = 0

    f = 0.75 * random.choice(l)

    t = 0

    while True:

        t += 1

        time.sleep(1)

        if f <= t:

            t_color = doss1.cell(row=1, column=1).value

            rcolor = colorl[t_color]

            t_color += 1

            doss1.cell(row=1, column=1).value = t_color

            if t_color == len(colorl):

                doss1.cell(row=1, column=1).value = 0
                
            while t_print < 75:
                
                print(colored("########################################################################################################################################################################################################################################", rcolor))
                
                t_print += 1
                
            t_print = 0
                       
            deb = time.time()

            while True:

                if keyboard.read_key() == "b":

                    end = time.time()

                    print(end - deb, "s")

                    q = str(input("Continue? (y/n)"))

                    rl.append(end - deb)

                    if q == "y":

                        fun()

                    else:

                        doss.save("reflex.xlsx")

                        t = 0

                        moy = 1

                        while t < len(rl):

                            moy = rl[t] + moy

                            plt.plot([t], [rl[t]], color="red", marker="*")

                            t += 1

                        moy = moy / len(rl)

                        print("The average is", moy, "s")

                        plt.savefig(namesaved)
                        
                        plt.show()

                        sys.exit()

fun()


