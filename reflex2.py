#!usr/bin python3

import time

import os

import matplotlib.pyplot as plt

import sys

import random

import keyboard

from termcolor import colored

from openpyxl import load_workbook, Workbook

import math

doss = load_workbook("r2.xlsx")

doss1 = doss.active

numd = doss1.cell(row=1, column=2).value + 1

doss1.cell(row=1, column=2).value = numd

namesaved = "refllex_evolved" + str(numd) + ".jpg"

l = []

rl = []

keyboardl = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t",
             "u", "v", "w", "x", "y", "z"]

def fun():

    up = 0

    t = 0

    wrong = 0

    while t < 20:

        os.system("clear")

        lechoix = random.choice(keyboardl)

        print(colored(lechoix, "red"))

        deb = time.time()

        while up != 1:

            inpt = keyboard.read_key()

            if inpt != None:
                
                end = time.time()
                
                if inpt != lechoix:

                    wrong += 1

                rl.append(end - deb)

                up = 1

        up = 0

        t += 1

    os.system("clear")

    cont = str(input("Do you want to continue? (y/n)"))

    if cont == "y":

        fun()

    else:

        t = 0

        moy = 1

        while t + 1 < len(rl):

            moy += rl[t]

            plt.plot([t,t + 1], [rl[t], rl[t + 1]], color="red", marker="*")

            t += 1

        moy = moy / len(rl)

        print("Average reaction time is", moy, "with", wrong, "errors")

        saved = str(input("Do you want to save results? (y/n)"))

        plt.ylabel("reaction time (s)")

        plt.xlabel("attempt")

        if saved == "y":

            doss.save("r2.xlsx")

            plt.savefig(namesaved)

        sg = str(input("Show graph? (y/n)"))

        if sg == "y":

            plt.show()

        sys.exit()

fun()

        
